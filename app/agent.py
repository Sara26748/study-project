import re
from pypdf.errors import PdfReadError
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, abort
from flask_login import login_required, current_user
from . import db
from .models import Project, Requirement, RequirementVersion, version_label
from .services.ai_client import generate_requirements

agent_bp = Blueprint('agent', __name__, template_folder='templates/agent')
output_language = None


def check_project_access(project):
    """Check if current user has access to the project (owner or shared)."""
    if project.user_id != current_user.id and current_user not in project.shared_with:
        abort(403)

def has_quantifiable_signal(title: str, description: str) -> bool:
    """Heuristic to detect quantifiable requirements from text signals."""
    text = " ".join([title or "", description or ""]).lower()
    if not text.strip():
        return False

    comparator_pattern = r"(<=|>=|<|>|=)"
    number_pattern = r"\d+(?:[.,]\d+)?"
    unit_pattern = (
        r"(ms|s|sec|sek|seconds|minutes|min|h|hz|khz|mhz|ghz|kb|mb|gb|tb|"
        r"bps|kbps|mbps|gbps|mb/s|gb/s|%|°c|c|w|kw|v|a|mah|rpm|fps|km/h|m/s|liter|l|g|kg|km|m|mm)"
    )
    keyword_pattern = r"(max|maximum|min|minimum|höchstens|mindestens|weniger als|mehr als)"

    # Standardmuster
    if re.search(rf"{comparator_pattern}\s*{number_pattern}", text):
        return True
    if re.search(rf"{number_pattern}\s*{unit_pattern}", text):
        return True
    if re.search(rf"{keyword_pattern}\s*{number_pattern}", text):
        return True

    # Erweiterung: Muster wie '6 Liter pro 100 km', '5 l/100km', 'g/km', etc.
    extended_pattern = r"(\d+(?:[.,]\d+)?\s*(liter|l|g|kg|km|m|mm)\s*(pro|/|je)\s*\d+(?:[.,]\d+)?\s*(km|m|h|min|sek|s|stunde|tag|jahr|monate|wochen))"
    if re.search(extended_pattern, text):
        return True

    return False

def normalize_key(title: str) -> str:
    """Creates a stable, lowercase key from a title string."""
    if not title:
        return ""
    s = title.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def next_version_info(req: Requirement) -> tuple[int, str]:
    """Determines the next version index and label for a requirement."""
    if not req.versions:
        return 1, version_label(1)
    # The versions are ordered by version_index, so the last one is the latest.
    last_idx = req.versions[-1].version_index
    new_idx = last_idx + 1
    return new_idx, version_label(new_idx)


@agent_bp.route('/agent/<int:project_id>', methods=['GET'])
@login_required
def agent_page(project_id):
    """
    Render the AI agent page for a specific project.
    Only accessible by the project owner.
    """
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    return render_template('agent.html', project=project)


@agent_bp.route('/agent/generate/<int:project_id>', methods=['POST'])
@login_required
def generate(project_id):
    """
    Generate requirements using AI and create versioned entries in the database.
    """
    project = Project.query.get_or_404(project_id)
    # Check access (owner or shared)
    if project.user_id != current_user.id and current_user not in project.shared_with:
        return jsonify({'ok': False, 'error': 'Zugriff verweigert.'}), 403

    # Handle both JSON (legacy) and FormData
    user_description = ""
    inputs_dict = {}
    excel_context = ""
    pdf_context = ""
    product_system = ""
    ai_model = None
    num_requirements = None
    num_requirements_mode = None
    num_requirements_value = None
    output_language = None
    improve_only = False
    extend_existing = False
    has_pdf = False

    if request.content_type and 'application/json' in request.content_type:
        try:
            data = request.get_json()
            if not data:
                return jsonify({'ok': False, 'error': 'Keine Daten empfangen.'}), 400

            user_description = data.get('user_description', '').strip() or None
            inputs_array = data.get('inputs', [])
            inputs_dict = {item.get('key'): item.get('value') for item in inputs_array if item.get('key')}
            product_system = data.get('product_system', '').strip()
            ai_model = data.get('ai_model', '').strip() or None
            improve_only = data.get('improve_only', False)
            extend_existing = data.get('extend_existing', False)
            output_language = data.get('output_language', '').strip() or None

            # Handle num_requirements_mode and num_requirements_value
            num_req_mode = data.get('num_requirements_mode', 'auto')
            if num_req_mode == 'manual':
                num_req_mode = 'exact'
            if num_req_mode in {'exact', 'min', 'max'}:
                try:
                    num_requirements_value = int(data.get('num_requirements_value', 5))
                except (ValueError, TypeError):
                    num_requirements_value = None
            num_requirements_mode = num_req_mode
            if num_req_mode == 'exact' and num_requirements_value:
                num_requirements = num_requirements_value
        except Exception:
            return jsonify({'ok': False, 'error': 'Ungültiges JSON-Format.'}), 400
    else:
        # Handle FormData
        user_description = request.form.get('user_description', '').strip() or None
        product_system = request.form.get('product_system', '').strip()
        ai_model = request.form.get('ai_model', '').strip() or None
        improve_only = request.form.get('improve_only', 'false') == 'true'
        extend_existing = request.form.get('extend_existing', 'false') == 'true'
        output_language = request.form.get('output_language', '').strip() or None

        # Handle num_requirements_mode and num_requirements_value
        num_req_mode = request.form.get('num_requirements_mode', 'auto')
        if num_req_mode == 'manual':
            num_req_mode = 'exact'
        if num_req_mode in {'exact', 'min', 'max'}:
            try:
                num_requirements_value = int(request.form.get('num_requirements_value', 5))
            except (ValueError, TypeError):
                num_requirements_value = None
        num_requirements_mode = num_req_mode
        if num_req_mode == 'exact' and num_requirements_value:
            num_requirements = num_requirements_value

        keys = request.form.getlist('key[]')
        values = request.form.getlist('value[]')
        for k, v in zip(keys, values):
            if k and k.strip():
                inputs_dict[k.strip()] = v.strip()
        # Handle Excel file (only if not improving existing, or as supplemental context)
        if 'excel_file' in request.files:
            file = request.files['excel_file']
            if file and file.filename != '' and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file, data_only=True)
                    ws = wb.active
                    
                    excel_context = "\n\n--- KONTEXT AUS EXCEL-DATEI ---\n"
                    # Read max 50 rows to avoid context overflow
                    count = 0
                    headers = []
                    
                    for row in ws.iter_rows(values_only=True):
                        if count == 0:
                            headers = [str(cell) if cell else "" for cell in row]
                            excel_context += " | ".join(headers) + "\n"
                            count += 1
                            continue
                        if count > 50:
                            excel_context += "... (weitere Zeilen aus Platzgrnden ausgelassen)\n"
                            break
                        row_vals = [str(cell) if cell is not None else "" for cell in row]
                        excel_context += " | ".join(row_vals) + "\n"
                        count += 1
                    excel_context += "--- ENDE EXCEL-KONTEXT ---\n"
                    if not improve_only:
                        excel_context += "\nWICHTIG: Die oben aufgefhrten Anforderungen aus der Excel-Datei sind BESTEHENDE Anforderungen. Du sollst:\n"
                        excel_context += "1. Diese bestehenden Anforderungen verbessern, aktualisieren und in deine Ausgabe aufnehmen\n"
                        excel_context += "2. Zusätzlich neue Anforderungen erstellen, die der User explizit anfordert (siehe Beschreibung oben)\n"
                        excel_context += "3. Weitere passende Anforderungen generieren, die zum Gesamtkontext passen\n"
                        excel_context += "Die bestehenden Anforderungen aus Excel drfen NICHT ignoriert werden!"
                except Exception as e:
                    print(f"Fehler beim Lesen der Excel-Datei: {e}")
                    # We continue without the excel content if it fails
        # Handle PDF file (as supplemental context)
        if 'pdf_file' in request.files:
            pdf_file = request.files['pdf_file']
            if pdf_file and pdf_file.filename != '':
                if not pdf_file.filename.lower().endswith('.pdf'):
                    return jsonify({'ok': False, 'error': 'Bitte lade eine gültige PDF-Datei hoch.'}), 400
                try:
                    from io import BytesIO
                    from pypdf import PdfReader
                    pdf_bytes = pdf_file.read()
                    if not pdf_bytes:
                        return jsonify({'ok': False, 'error': 'Die PDF-Datei ist leer oder konnte nicht gelesen werden.'}), 400
                    reader = PdfReader(BytesIO(pdf_bytes))
                    if reader.is_encrypted:
                        try:
                            reader.decrypt("")
                        except Exception:
                            return jsonify({'ok': False, 'error': 'Die PDF-Datei ist verschlüsselt und kann nicht gelesen werden.'}), 400
                    extracted_pages = []
                    for page in reader.pages:
                        page_text = page.extract_text() or ""
                        if page_text.strip():
                            extracted_pages.append(page_text.strip())
                    pdf_text = "\n\n".join(extracted_pages).strip()
                    if not pdf_text:
                        return jsonify({'ok': False, 'error': 'Die PDF-Datei konnte nicht gelesen werden (kein Text gefunden).'}), 400
                    max_chars = 15000
                    if len(pdf_text) > max_chars:
                        pdf_text = pdf_text[:max_chars].rstrip()
                        pdf_text += "\n... (PDF-Text aus Platzgründen gekürzt)"
                    pdf_context = "\n\n--- KONTEXT AUS PDF (PFLICHTENHEFT) ---\n"
                    pdf_context += pdf_text
                    pdf_context += "\n--- ENDE PDF-KONTEXT ---\n"
                    has_pdf = True
                except PdfReadError as e:
                    print(f"Fehler beim Lesen der PDF-Datei: {e}")
                    return jsonify({'ok': False, 'error': 'Die PDF-Datei ist beschädigt oder hat ein nicht unterstütztes Format.'}), 400
                except Exception as e:
                    print(f"Fehler beim Lesen der PDF-Datei: {e}")
                    return jsonify({'ok': False, 'error': 'Die PDF-Datei konnte nicht gelesen werden.'}), 400

    # Auto-detect custom columns from Excel headers if any were found
    if 'headers' in locals() and headers:
        try:
            current_custom_columns = project.get_custom_columns()
            # Normalize current columns for checking existence (lowercase)
            current_custom_lower = [c.lower() for c in current_custom_columns]
            
            # Standard columns to ignore
            standard_columns = ['title', 'titel', 'description', 'beschreibung', 'category', 'kategorie', 'status', 'id', 'version', 'req_id', 'req-id']
            
            new_columns_found = False
            for header in headers:
                if not header:
                    continue
                    
                h_str = str(header).strip()
                if not h_str:
                    continue
                
                # Check if it's a new custom column
                if h_str.lower() not in standard_columns and h_str.lower() not in current_custom_lower:
                    current_custom_columns.append(h_str)
                    current_custom_lower.append(h_str.lower())
                    new_columns_found = True
            
            if new_columns_found:
                project.set_custom_columns(current_custom_columns)
                db.session.commit()
                # Flash message is not possible here as it's an AJAX request, 
                # but the UI will show the new columns on reload.
        except Exception as e:
            print(f"Error updating custom columns: {e}")


    # Append Excel context to user description if present
    full_description = user_description or ""
    has_excel = False
    if excel_context:
        full_description += excel_context
        has_excel = True
    if pdf_context:
        full_description += pdf_context

    # Get project's custom columns
    custom_columns = project.get_custom_columns()
    
    # Handle custom columns from form (if provided)
    if 'custom_columns' in request.form:
        try:
            import json
            form_custom_columns = json.loads(request.form.get('custom_columns', '[]'))
            
            # Merge with existing custom columns (avoid duplicates)
            current_custom_lower = [c.lower() for c in custom_columns]
            for col in form_custom_columns:
                if col and col.strip() and col.lower() not in current_custom_lower:
                    custom_columns.append(col.strip())
                    current_custom_lower.append(col.lower())
            
            # Update project with new columns
            if form_custom_columns:
                project.set_custom_columns(custom_columns)
                db.session.commit()
        except Exception as e:
            print(f"Error processing custom columns from form: {e}")
    
    # Build complete columns list: title, description, custom columns, category, status
    columns = ["title", "description"] + custom_columns + ["category", "status"]

    # If Improve Existing Mode
    if improve_only:
        # Fetch existing requirements
        existing_reqs = Requirement.query.filter_by(project_id=project_id).all()
        if not existing_reqs:
            return jsonify({'ok': False, 'error': 'Keine bestehenden Anforderungen gefunden, die verbessert werden können.'}), 400
            
        req_context = "\n\n--- BESTEHENDE ANFORDERUNGEN ZUR VERBESSERUNG ---\n"
        req_count = 0
        
        # Add ID to columns for AI to return it
        columns = ["id"] + columns
        
        for req in existing_reqs:
            latest = req.get_latest_version()
            if not latest:
                continue
                
            req_context += f"ID: {req.id}\n"
            req_context += f"Titel: {latest.title}\n"
            req_context += f"Beschreibung: {latest.description}\n"
            req_context += f"Kategorie: {latest.category}\n"
            
            # Add custom data
            custom = latest.get_custom_data()
            if custom:
                for k, v in custom.items():
                    req_context += f"{k}: {v}\n"
            
            req_context += "---\n"
            req_count += 1
            
        full_description += req_context
        # Override num_requirements to match exact count
        num_requirements = req_count
    
    # If Extend Existing Mode
    elif extend_existing:
        # Fetch existing requirements for context
        existing_reqs = Requirement.query.filter_by(project_id=project_id).all()
        if existing_reqs:
            req_context = "\n\n--- BESTEHENDE PROJEKT-ANFORDERUNGEN (NICHT VERÄNDERN, NUR ERGÄNZEN) ---\n"
            
            for req in existing_reqs:
                latest = req.get_latest_version()
                if not latest:
                    continue
                    
                req_context += f"- ID {req.id}: {latest.title} ({latest.description})\n"
                
            req_context += "--- ENDE BESTEHENDE ANFORDERUNGEN ---\n"
            full_description += req_context

    try:
        requirements_data = generate_requirements(
            full_description,
            inputs_dict,
            columns,
            ai_model=ai_model,
            num_requirements=num_requirements,
            num_requirements_mode=num_requirements_mode,
            num_requirements_value=num_requirements_value,
            product_system=product_system,
            has_excel_context=has_excel,
            has_pdf_context=has_pdf,
            improve_only=improve_only,
            extend_existing=extend_existing,
            output_language=output_language
        )
        saved_count = 0
        for item in requirements_data:
            title = item.get("title", "").strip()
            if not title:
                continue  # Skip requirements without a title

            description = item.get("description", "").strip()
            category = item.get("category", "") or ""
            status = item.get("status", "") or "Offen"
            req = None
            # Check for ID first (especially for improve_only mode)
            req_id_val = item.get("id")
            if req_id_val:
                try:
                    r_id = int(str(req_id_val).strip())
                    req = Requirement.query.get(r_id)
                    if req and req.project_id != project_id:
                        req = None # Security check
                except (ValueError, TypeError):
                    pass

            if not req:
                from .agent import normalize_key
                key = normalize_key(title)
                # Find existing logical requirement in the current project by key
                req = Requirement.query.filter_by(project_id=project_id, key=key).first()

                if not req:
                    # It's a new logical requirement, create it
                    req = Requirement(project_id=project_id, key=key)
                    db.session.add(req)
                    # Flush to get the ID for the foreign key relationship
                    db.session.flush()
                    version_index, label = 1, version_label(1)
                else:
                    # It's a new version of an existing requirement
                    version_index, label = next_version_info(req)
            else:
                 # Found by ID
                 version_index, label = next_version_info(req)

            # Create the new version
            new_version = RequirementVersion(
                requirement_id=req.id,
                version_index=version_index,
                version_label=label,
                title=title,
                description=description,
                category=category,
                status=status,
                created_by_id=current_user.id  # Track who created this version
            )
            
            # Save custom column data
            custom_data = {}
            for col in custom_columns:
                value = item.get(col, "")
                if value:
                    custom_data[col] = value
            
            # Handle is_quantifiable from AI
            is_quantifiable = item.get("is_quantifiable", False)
            if isinstance(is_quantifiable, bool):
                custom_data['is_quantifiable'] = 'true' if is_quantifiable else 'false'
            elif isinstance(is_quantifiable, str):
                custom_data['is_quantifiable'] = 'true' if is_quantifiable.lower() in ['true', '1', 'yes'] else 'false'
            else:
                custom_data['is_quantifiable'] = 'false'
            

            # Heuristic: If is_quantifiable is 'false' but signals are present, set to 'true'
            if custom_data.get('is_quantifiable') == 'false' and has_quantifiable_signal(title, description):
                custom_data['is_quantifiable'] = 'true'

            new_version.set_custom_data(custom_data)
            
            db.session.add(new_version)
            db.session.flush()  # Get the ID for history entry
            
            # Create history entry for creation
            from .models import RequirementVersionHistory
            import json
            history_entry = RequirementVersionHistory(
                version_id=new_version.id,
                changed_by_id=current_user.id,
                change_type='created',
                changes=json.dumps({'action': 'Version erstellt', 'version': label})
            )
            db.session.add(history_entry)
            
            # Setze funktional automatisch beim Generieren vor RequirementVersion-Erstellung.
            funktional_value = None
            raw_funktional = item.get("funktional", None)
            if isinstance(raw_funktional, bool):
                funktional_value = raw_funktional
            elif isinstance(raw_funktional, str):
                if raw_funktional.strip().lower() in ["true", "1", "yes", "ja"]:
                    funktional_value = True
                elif raw_funktional.strip().lower() in ["false", "0", "no", "nein"]:
                    funktional_value = False

            if funktional_value is None and category:
                category_lower = category.strip().lower()
                if "nicht" in category_lower and "funktional" in category_lower:
                    funktional_value = False
                elif "funktional" in category_lower:
                    funktional_value = True

            if funktional_value is not None:
                req.funktional = funktional_value

            saved_count += 1

        db.session.commit()
        
        # Create notifications for newly created requirements
        # Note: This is done after commit to ensure all versions are saved
        try:
            from .utils.notifications import notify_requirement_created
            from datetime import timedelta, datetime as dt
            for req in Requirement.query.filter_by(project_id=project_id).all():
                latest_version = req.get_latest_version()
                if latest_version and latest_version.created_by_id == current_user.id:
                    # Check if this version was just created (within last 2 seconds)
                    if latest_version.created_at > dt.utcnow() - timedelta(seconds=2):
                        notify_requirement_created(latest_version, current_user)
        except Exception:
            # Don't fail generation if notification fails
            pass

        return jsonify({
            'ok': True,
            'count': saved_count,
            'redirect': url_for('main.manage_project', project_id=project_id),
            'was_improved': improve_only,
            'was_extended': extend_existing
        }), 200

    except ValueError as e:
        return jsonify({'ok': False, 'error': f'Konfigurationsfehler: {str(e)}'}), 500
    except RuntimeError as e:
        return jsonify({'ok': False, 'error': f'KI-Service-Fehler: {str(e)}'}), 500
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'error': f'Ein unerwarteter Fehler ist aufgetreten: {str(e)}'}), 500
