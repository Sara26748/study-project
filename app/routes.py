from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, jsonify, session, current_app, send_file
from flask_login import login_required, current_user
from sqlalchemy import func, and_
from datetime import datetime
from urllib.parse import urlparse
import json
from . import db
from .models import Project, Requirement, RequirementVersion, RequirementComment, Notification, User
from .services.ai_client import generate_requirements

bp = Blueprint('main', __name__)


def _int_to_roman(num: int) -> str:
    """Convert integer to Roman numerals (supports positive numbers)."""
    if num <= 0:
        return ""
    mapping = [
        (1000, "M"),
        (900, "CM"),
        (500, "D"),
        (400, "CD"),
        (100, "C"),
        (90, "XC"),
        (50, "L"),
        (40, "XL"),
        (10, "X"),
        (9, "IX"),
        (5, "V"),
        (4, "IV"),
        (1, "I"),
    ]
    result = ""
    for value, numeral in mapping:
        while num >= value:
            result += numeral
            num -= value
    return result.lower()


def _roman_to_int(value: str) -> int:
    """Convert Roman numerals to integer; returns 0 if invalid/empty."""
    if not value:
        return 0
    mapping = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    total = 0
    prev = 0
    for ch in reversed(value.upper()):
        current = mapping.get(ch, 0)
        if current < prev:
            total -= current
        else:
            total += current
            prev = current
    return total


def _parse_revision_number(value: str) -> int:
    """Parse revision label like 'ii' or legacy 'ii.A' into a number."""
    if not value or value == 'Entwurf':
        return 0
    raw = value.strip()
    roman_part = raw.split('.', 1)[0].strip() if '.' in raw else raw
    return _roman_to_int(roman_part)


def _parse_revision_round(value: str) -> str:
    """Parse legacy revision label like 'ii.A' into round letter 'A'."""
    if not value or value == 'Entwurf':
        return 'A'
    raw = value.strip()
    if '.' in raw:
        _, round_part = raw.split('.', 1)
        return round_part.strip().upper()[:1] or 'A'
    return 'A'


def _format_revision_label(number: int) -> str:
    """Format revision label like 'i'."""
    if number <= 0:
        return ''
    return _int_to_roman(number)


def _next_round_letter(letter: str) -> str:
    """Advance revision round letter (A -> B)."""
    base = (letter or 'A').upper()
    if len(base) != 1 or not ('A' <= base <= 'Z'):
        return 'A'
    if base == 'Z':
        return 'Z'
    return chr(ord(base) + 1)


def _parse_revision_key(value: str) -> tuple[str, int]:
    """Parse revision key like 'A-2' into ('A', 2)."""
    if not value:
        return '', 0
    raw = value.strip()
    if '-' in raw:
        round_part, num_part = raw.split('-', 1)
        round_letter = round_part.strip().upper()[:1] or 'A'
        try:
            number = int(num_part.strip())
        except ValueError:
            number = 0
        return round_letter, number
    try:
        number = int(raw)
    except ValueError:
        number = 0
    return 'A', number


def _normalize_revision_label(value: str) -> str:
    """Normalize revision labels for display."""
    number = _parse_revision_number(value)
    return _format_revision_label(number)

# Route: Manuelle Anforderung erstellen
@bp.route("/project/<int:project_id>/manual_requirement", methods=["POST"])
@login_required
def create_manual_requirement(project_id):
    project = Project.query.get_or_404(project_id)
    check_project_access(project)

    title = request.form.get("title", "").strip()
    description = request.form.get("description", "").strip()
    category = request.form.get("category", "").strip()
    status = request.form.get("status", "Entwurf").strip()
    is_quantifiable = request.form.get("is_quantifiable") == "on"
    funktional = request.form.get("funktional") == "on"

    if not title or not description:
        flash("Name und Beschreibung sind erforderlich.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))

    # Key für Duplikatserkennung
    from .agent import normalize_key
    key = normalize_key(title)
    req = Requirement.query.filter_by(project_id=project_id, key=key).first()
    if not req:
        req = Requirement(project_id=project_id, key=key, funktional=funktional)
        db.session.add(req)
        db.session.flush()
        version_index = 1
        version_label = 'A'
    else:
        # Update funktional, falls geändert
        if req.funktional != funktional:
            req.funktional = funktional
        last_version = req.versions[-1] if req.versions else None
        version_index = last_version.version_index + 1 if last_version else 1
        version_label = chr(ord('A') + (version_index - 1))

    new_version = RequirementVersion(
        requirement_id=req.id,
        version_index=version_index,
        version_label=version_label,
        title=title,
        description=description,
        category=category,
        status=status,
        revision=None,
        created_by_id=current_user.id
    )
    # Custom Data
    custom_data = {}
    if is_quantifiable:
        custom_data['is_quantifiable'] = 'true'
    else:
        custom_data['is_quantifiable'] = 'false'
    # Dynamische Spalten
    for col in project.get_custom_columns():
        val = request.form.get(f'custom_{col}', '').strip()
        if val:
            custom_data[col] = val
    if custom_data:
        new_version.set_custom_data(custom_data)
    db.session.add(new_version)
    db.session.flush()
    # History
    from .models import RequirementVersionHistory
    import json
    history_entry = RequirementVersionHistory(
        version_id=new_version.id,
        changed_by_id=current_user.id,
        change_type='created',
        changes=json.dumps({'action': 'Manuell erstellt', 'version': version_label})
    )
    db.session.add(history_entry)
    db.session.commit()
    # Notification
    try:
        notify_requirement_created(new_version, current_user)
    except Exception:
        pass
    flash(f"Anforderung '{title}' wurde erfolgreich erstellt.", "success")
    return redirect(url_for('main.manage_project', project_id=project_id))

def check_project_access(project):
    """Check if current user has access to the project (owner or shared)."""
    if project.user_id != current_user.id and current_user not in project.shared_with:
        abort(403)

def check_requirement_access(requirement):
    """Check if current user has access to the requirement's project (owner or shared)."""
    check_project_access(requirement.project)

def check_version_access(version):
    """Check if current user has access to the version's requirement project (owner or shared)."""
    check_project_access(version.requirement.project)

@bp.route("/")
@login_required
def home():
    # Get projects owned by the user
    owned_projects = Project.query.filter_by(user_id=current_user.id).all()
    
    # Get projects shared with the user
    shared_projects = current_user.shared_projects.all()
    
    # Combine both lists (owned first, then shared)
    projects = owned_projects + shared_projects
    
    return render_template("start.html", projects=projects)


@bp.route("/set_language/<lang>")
@login_required
def set_language(lang):
    supported = current_app.config.get("BABEL_SUPPORTED_LOCALES", [])
    if lang in supported:
        session["lang"] = lang

    referrer = request.referrer or url_for("main.home")
    try:
        parsed = urlparse(referrer)
        if parsed.netloc and parsed.netloc != request.host:
            return redirect(url_for("main.home"))
    except Exception:
        return redirect(url_for("main.home"))

    return redirect(referrer)


@bp.route("/project/<int:project_id>/mention_suggestions")
@login_required
def mention_suggestions(project_id):
    project = Project.query.get_or_404(project_id)
    check_project_access(project)

    q = (request.args.get("q", "") or "").strip().lower()

    # Kandidaten: Owner + shared_with + current_user
    candidates = []

    owner = User.query.get(project.user_id)
    if owner:
        candidates.append(owner)

    for u in project.shared_with:
        candidates.append(u)

    candidates.append(current_user)

    # Duplikate entfernen
    uniq = {}
    for u in candidates:
        uniq[u.id] = u

    results = []
    for u in uniq.values():
        email = u.email or ""
        username = (email.split("@")[0] if email else "").strip()

        # Filter: query passt in username oder email
        if not q or q in username.lower() or q in email.lower():
            results.append({
                "id": u.id,
                "username": username,
                "email": email
            })

    results = sorted(results, key=lambda x: (x["username"] or x["email"]))[:8]
    return jsonify({"users": results})

@bp.route("/create", methods=['GET', 'POST'])
@login_required
def create():
    if request.method == 'POST':
        project_name = request.form.get('project_name')
        if project_name:
            # The concept of dynamic columns is removed in the new model.
            new_project = Project(name=project_name, user_id=current_user.id)
            db.session.add(new_project)
            db.session.commit()
        return redirect(url_for('main.home'))
    # The create.html is now a generic "new project" page if no project is passed.
    return render_template("create.html", project=None)

@bp.route("/project/<int:project_id>")
@login_required
def manage_project(project_id):
    project = Project.query.get_or_404(project_id)
    # Check if user is owner or has shared access
    if project.user_id != current_user.id and current_user not in project.shared_with:
        abort(403)

    session["last_project_id"] = project_id

    # Get all requirements with ALL versions (not just the latest)
    # Filter out deleted requirements
    requirements = (
        Requirement.query
        .filter_by(project_id=project_id, is_deleted=False)
        .all()
    )
    
    # For each requirement, get all versions
    req_with_versions = []
    for req in requirements:
        # Get all versions for this requirement
        versions = req.versions
        if versions:  # Only include requirements that have versions
            req_with_versions.append((req, versions))
    
    # Get custom columns for this project
    custom_columns = project.get_custom_columns()
    
    selected_version_id = request.args.get("selected_version_id", type=int)

    return render_template(
        "create.html", 
        project=project, 
        req_with_versions=req_with_versions,
        custom_columns=custom_columns,
        selected_version_id=selected_version_id
    )

@bp.route("/deleted_requirements")
@login_required
def deleted_requirements_overview():
    """Show all deleted requirements across all user's projects."""
    owned_projects = Project.query.filter_by(user_id=current_user.id).all()
    shared_projects = current_user.shared_projects.all()
    projects_by_id = {project.id: project for project in owned_projects}
    for project in shared_projects:
        projects_by_id.setdefault(project.id, project)

    # Collect deleted requirements from all projects
    all_deleted = []
    for project in projects_by_id.values():
        deleted_reqs = Requirement.query.filter_by(
            project_id=project.id, 
            is_deleted=True
        ).all()
        
        for req in deleted_reqs:
            latest_version = req.get_latest_version()
            if latest_version:
                all_deleted.append({
                    'project': project,
                    'requirement': req,
                    'version': latest_version
                })
    
    last_project_id = session.get("last_project_id")
    if last_project_id not in projects_by_id:
        last_project_id = None
    if not last_project_id:
        last_project_id = next(iter(projects_by_id), None)
    if last_project_id:
        session["last_project_id"] = last_project_id

    return render_template(
        "deleted_requirements_overview.html",
        deleted_items=all_deleted,
        last_project_id=last_project_id
    )

@bp.route("/requirement/<int:rid>/history")
@login_required
def requirement_history(rid):
    from .models import RequirementVersionHistory
    
    req = Requirement.query.get_or_404(rid)
    # Authorization check: ensure the user has access to the project (owner or shared)
    check_requirement_access(req)
    
    # Get latest version
    latest_version = req.get_latest_version()
    if not latest_version:
        flash("No versions found for this requirement.", "warning")
        return redirect(url_for('main.manage_project', project_id=req.project_id))
    
    # Get all history entries for this version (ordered by date)
    history_entries = RequirementVersionHistory.query.filter_by(
        version_id=latest_version.id
    ).order_by(RequirementVersionHistory.created_at.asc()).all()
    
    # Build timeline: creation + all modifications
    timeline = []
    
    # Add creation entry
    if latest_version.created_by:
        timeline.append({
            'type': 'created',
            'user': latest_version.created_by,
            'timestamp': latest_version.created_at,
            'changes': {'action': 'Version erstellt'}
        })
    
    # Add all modification entries
    for entry in history_entries:
        timeline.append({
            'type': entry.change_type,
            'user': entry.changed_by,
            'timestamp': entry.created_at,
            'changes': entry.get_changes()
        })
    
    return render_template("requirement_history.html", 
                         req=req, 
                         version=latest_version,
                         timeline=timeline)


@bp.route("/project/delete/<int:project_id>", methods=['POST'])
@login_required
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    # Remove any active sessions for this project first to avoid
    # SQLAlchemy attempting to nullify the foreign key on ActiveSession
    # (the `project_id` column is NOT NULL).
    from .models import ActiveSession
    try:
        # Use a bulk delete to remove any active session rows referencing
        # this project. This issues a direct DELETE statement.
        ActiveSession.query.filter_by(project_id=project_id).delete()
    except Exception:
        # If something goes wrong, continue and let the later delete/commit
        # raise the appropriate error so it can be debugged.
        pass

    db.session.delete(project)
    db.session.commit()
    flash(f"Projekt '{project.name}' wurde geloescht.", "success")
    return redirect(url_for('main.home'))


# The routes below are now obsolete due to the data model refactoring
# and have been removed:
# - /project/<int:project_id>/deleted
# - /deleted_requirements_overview
# - /move/<int:project_id>/<int:req_id>/<string:from_table>/<string:to_table>
# - /edit/<int:project_id>/<int:req_id>
# - /export/<int:project_id>/<string:format>
# - /delete_column/<int:project_id>
# - /delete_requirement_permanently/<int:project_id>/<int:req_id>
# The simple /requirements and /add_requirement routes were also based on the old model.

# Routes for dynamic columns
@bp.route("/project/<int:project_id>/add_column", methods=['POST'])
@login_required
def add_column(project_id):
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    
    column_name = request.form.get('column_name', '').strip()
    if not column_name:
        flash("Column name cannot be empty.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    # Get current columns and add the new one
    columns = project.get_custom_columns()
    if column_name in columns:
        flash(f"Column '{column_name}' already exists.", "warning")
    else:
        columns.append(column_name)
        project.set_custom_columns(columns)
        db.session.commit()
        flash(f"Column '{column_name}' added successfully.", "success")
    
    return redirect(url_for('main.manage_project', project_id=project_id))

@bp.route("/project/<int:project_id>/remove_column/<column_name>", methods=['POST'])
@login_required
def remove_column(project_id, column_name):
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    
    # Protected columns that cannot be deleted (ID and Version are database columns, not custom columns)
    PROTECTED_COLUMNS = ['title', 'description', 'category', 'status', 'titel', 'beschreibung', 'kategorie', 'id', 'version', 'ver', 'version_label', 'version_index']
    
    if column_name.lower() in [c.lower() for c in PROTECTED_COLUMNS]:
        flash(f"Die Spalte '{column_name}' ist geschützt und kann nicht gelöscht werden.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    # Get current columns and remove the specified one
    columns = project.get_custom_columns()
    if column_name in columns:
        columns.remove(column_name)
        project.set_custom_columns(columns)
        db.session.commit()
        flash(f"Column '{column_name}' removed successfully.", "success")
    else:
        flash(f"Column '{column_name}' not found.", "warning")
    
    return redirect(url_for('main.manage_project', project_id=project_id))

# Route to update custom column data for a requirement version
@bp.route("/requirement_version/<int:version_id>/update_custom_data", methods=['POST'])
@login_required
def update_custom_data(version_id):
    version = RequirementVersion.query.get_or_404(version_id)
    # Authorization check
    check_version_access(version)

    if version.status == 'Verworfen':
        return jsonify({'success': False, 'error': 'Requirement is rejected'}), 400

    # After a release, allow only revisions, not direct edits
    if version.requirement.has_release_event():
        flash("Diese Anforderung wurde bereits freigegeben und kann nur noch revidiert werden.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )
    
    column_name = request.form.get('column_name')
    value = request.form.get('value', '').strip()
    
    # Get current custom data and update it
    custom_data = version.get_custom_data()
    custom_data[column_name] = value
    version.set_custom_data(custom_data)
    db.session.commit()
    
    return jsonify({'success': True})

# Route to update requirement status
@bp.route("/requirement_version/<int:version_id>/update_status", methods=['POST'])
@login_required
def update_status(version_id):
    version = RequirementVersion.query.get_or_404(version_id)
    # Authorization check
    check_version_access(version)

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    status = request.form.get('status')
    if version.status == 'Verworfen' and status != 'Verworfen':
        if is_ajax:
             return jsonify({'success': False, 'error': 'Requirement is rejected'}), 400
        flash("Diese Anforderung ist verworfen und kann nur noch gelöscht werden.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )
    if status in ['Entwurf', 'In Bearbeitung', 'Freigabe', 'Verworfen']:
        version.status = status
        db.session.commit()
        
        if is_ajax:
             return jsonify({'success': True, 'status': status, 'color': version.get_status_color()})
             
        flash(f"Status updated to '{status}'.", "success")
    else:
        if is_ajax:
             return jsonify({'success': False, 'error': 'Invalid status'}), 400
        flash("Invalid status value.", "danger")
    
    return redirect(
        url_for(
            'main.manage_project',
            project_id=version.requirement.project_id,
            selected_version_id=version.id,
        )
    )


@bp.route("/project/<int:project_id>/kanban")
@login_required
def kanban_view(project_id):
    project = Project.query.get_or_404(project_id)
    check_project_access(project)

    # Get all active requirements
    requirements = Requirement.query.filter_by(project_id=project_id, is_deleted=False).all()
    
    # Sort into columns
    kanban_data = {
        'Entwurf': [],
        'In Bearbeitung': [],
        'Freigabe': []
    }
    
    for req in requirements:
        latest = req.get_latest_version()
        if latest and latest.status in kanban_data:
            kanban_data[latest.status].append({
                'req': req,
                'version': latest
            })
            
    custom_columns = project.get_custom_columns()
    
    return render_template(
        "kanban.html", 
        project=project, 
        kanban_data=kanban_data,
        custom_columns=custom_columns
    )


@bp.route("/project/<int:project_id>/requirements_status")
@login_required
def requirements_status_json(project_id):
    """API for live collaboration polling."""
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
        
    requirements = Requirement.query.filter_by(project_id=project_id, is_deleted=False).all()
    status_list = []
    
    for req in requirements:
        latest = req.get_latest_version()
        if latest:
            status_list.append({
                'req_id': req.id,
                'version_id': latest.id,
                'is_blocked': latest.is_blocked,
                'blocked_by': latest.blocked_by.email if latest.blocked_by else None,
                'status': latest.status
            })
            
    return jsonify(status_list)

@bp.route("/project/<int:project_id>/heartbeat", methods=['POST'])
@login_required
def project_heartbeat(project_id):
    """Update user's presence in the project."""
    from .models import ActiveSession
    from datetime import datetime, timedelta
    
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    
    # Find or create active session
    session = ActiveSession.query.filter_by(
        user_id=current_user.id,
        project_id=project_id
    ).first()
    
    if session:
        session.last_seen = datetime.utcnow()
    else:
        session = ActiveSession(
            user_id=current_user.id,
            project_id=project_id,
            last_seen=datetime.utcnow()
        )
        db.session.add(session)
    
    db.session.commit()
    
    # Clean up old sessions (older than 30 seconds)
    threshold = datetime.utcnow() - timedelta(seconds=30)
    ActiveSession.query.filter(ActiveSession.last_seen < threshold).delete()
    db.session.commit()
    
    return jsonify({'ok': True})

@bp.route("/project/<int:project_id>/active_users")
@login_required
def active_users(project_id):
    """Get list of currently active users in the project."""
    from .models import ActiveSession, User
    from datetime import datetime, timedelta
    
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    
    # Get sessions active in last 30 seconds
    threshold = datetime.utcnow() - timedelta(seconds=30)
    active_sessions = ActiveSession.query.filter(
        ActiveSession.project_id == project_id,
        ActiveSession.last_seen >= threshold
    ).all()
    
    users_data = []
    for session in active_sessions:
        if session.user_id != current_user.id:  # Don't include current user
            users_data.append({
                'id': session.user.id,
                'email': session.user.email,
                'initials': ''.join([word[0].upper() for word in session.user.email.split('@')[0].split('.')[:2]])
            })
    
    return jsonify(users_data)


# AJAX route to get all versions of a requirement
@bp.route("/requirement/<int:req_id>/versions_json")
@login_required
def requirement_versions_json(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    check_requirement_access(req)
    
    versions_data = []
    for ver in req.versions:
        versions_data.append({
            'id': ver.id,
            'version_index': ver.version_index,
            'version_label': ver.version_label,
            'title': ver.title,
            'description': ver.description,
            'category': ver.category,
            'status': ver.status,
            'status_color': ver.get_status_color(),
            'custom_data': ver.get_custom_data(),
            'created_at': ver.created_at.strftime('%Y-%m-%d %H:%M')
        })
    
    return jsonify(versions_data)


# AJAX route to get all revisions (snapshots) for a requirement
@bp.route("/requirement/<int:req_id>/revisions_json")
@login_required
def requirement_revisions_json(req_id):
    """Return revision snapshots scoped to a specific version (if provided)."""
    from .models import RequirementVersionHistory, RequirementVersion

    req = Requirement.query.get_or_404(req_id)
    check_requirement_access(req)

    version_id = request.args.get("version_id", type=int)

    # If no version_id provided, default to latest version to avoid mixing versions
    if version_id is None:
        latest = req.get_latest_version()
        version_id = latest.id if latest else None

    if not version_id:
        return jsonify([])

    version_obj = RequirementVersion.query.get_or_404(version_id)
    if version_obj.requirement_id != req.id:
        abort(404)

    version_labels = [ver.version_label for ver in (req.versions or [])]
    if not version_labels:
        version_labels = [version_obj.version_label]

    version_order = {label: idx for idx, label in enumerate(version_labels, start=1)}

    def normalize_round_label(value: str) -> str:
        candidate = (value or '').strip().upper()[:1]
        if not candidate:
            return version_labels[0] if version_labels else 'A'
        if candidate in version_order:
            return candidate
        return version_labels[0] if version_labels else candidate

    history_entries = (
        RequirementVersionHistory.query
        .filter(
            RequirementVersionHistory.version_id == version_obj.id,
            RequirementVersionHistory.change_type == 'revised'
        )
        .order_by(RequirementVersionHistory.created_at.asc())
        .all()
    )

    revisions_data = []
    baseline_added = False

    for entry in history_entries:
        try:
            changes = entry.get_changes()
        except Exception:
            changes = {}

        baseline_snapshot = changes.get('revision_baseline')
        if baseline_snapshot and not baseline_added:
            baseline_label = _normalize_revision_label(baseline_snapshot.get('revision_label'))
            baseline_number = baseline_snapshot.get('revision_number') or _parse_revision_number(baseline_label)
            baseline_round = normalize_round_label(
                baseline_snapshot.get('version_label') or _parse_revision_round(baseline_snapshot.get('revision_label'))
            )
            revisions_data.append({
                'revision_label': baseline_label,
                'revision_number': baseline_number or 0,
                'revision_key': f"{baseline_round}-{baseline_number or 0}",
                'version_label': baseline_round,
                'revision_sort': (version_order.get(baseline_round, 0) * 10) + (baseline_number or 0),
                'title': baseline_snapshot.get('title'),
                'description': baseline_snapshot.get('description'),
                'category': baseline_snapshot.get('category'),
                'status': baseline_snapshot.get('status'),
                'status_color': baseline_snapshot.get('status_color') or 'secondary',
                'custom_data': baseline_snapshot.get('custom_data') or {},
                'is_quantifiable': baseline_snapshot.get('is_quantifiable', False),
                'created_at': entry.created_at.isoformat()
            })
            baseline_added = True

    # Add fallback baseline when no baseline snapshot exists yet
    if not baseline_added:
        fallback_label = _normalize_revision_label(version_obj.revision)
        fallback_number = _parse_revision_number(fallback_label)
        fallback_round = normalize_round_label(version_obj.version_label)
        revisions_data.append({
            'revision_label': fallback_label,
            'revision_number': fallback_number,
            'revision_key': f"{fallback_round}-{fallback_number}",
            'version_label': fallback_round,
            'revision_sort': (version_order.get(fallback_round, 0) * 10) + (fallback_number or 0),
            'title': version_obj.title,
            'description': version_obj.description,
            'category': version_obj.category,
            'status': version_obj.status,
            'status_color': version_obj.get_status_color(),
            'custom_data': version_obj.get_custom_data(),
            'is_quantifiable': version_obj.get_custom_data().get('is_quantifiable') in ['true', True],
            'created_at': version_obj.created_at.isoformat()
        })

    for entry in history_entries:
        try:
            changes = entry.get_changes()
        except Exception:
            changes = {}

        snapshot = changes.get('revision_snapshot') or {}
        if not snapshot:
            continue

        snapshot_label = _normalize_revision_label(snapshot.get('revision_label'))
        snapshot_number = snapshot.get('revision_number') or _parse_revision_number(snapshot_label)
        snapshot_round = normalize_round_label(
            snapshot.get('version_label') or _parse_revision_round(snapshot.get('revision_label'))
        )

        revisions_data.append({
            'revision_label': snapshot_label,
            'revision_number': snapshot_number,
            'revision_key': f"{snapshot_round}-{snapshot_number or 0}",
            'version_label': snapshot_round,
            'revision_sort': (version_order.get(snapshot_round, 0) * 10) + (snapshot_number or 0),
            'title': snapshot.get('title'),
            'description': snapshot.get('description'),
            'category': snapshot.get('category'),
            'status': snapshot.get('status'),
            'status_color': snapshot.get('status_color') or 'secondary',
            'custom_data': snapshot.get('custom_data') or {},
            'is_quantifiable': snapshot.get('is_quantifiable', False),
            'created_at': entry.created_at.isoformat()
        })

    return jsonify(revisions_data)

# Route to update requirement version data
@bp.route("/requirement_version/<int:version_id>/update", methods=['POST'])
@login_required
def update_requirement_version(version_id):
    from .models import RequirementVersionHistory
    import json
    
    version = RequirementVersion.query.get_or_404(version_id)
    # Authorization check
    check_version_access(version)

    if version.status == 'Verworfen':
        flash("Diese Anforderung ist verworfen und kann nur noch gelöscht werden.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )
    
    # Get form data
    title = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    category = request.form.get('category', '').strip()

    # Validate required fields
    if not title or not description:
        flash("Title and description are required.", "danger")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )
    
    # Track changes for history (relative to the edited version)
    changes = {}
    if version.title != title:
        changes['title'] = f"{version.title} → {title}"
    if version.description != description:
        changes['description'] = "Beschreibung geändert"
    if (version.category or '') != category:
        changes['category'] = f"{version.category or '–'} → {category or '–'}"

    # Build custom data for the new version
    old_custom_data = version.get_custom_data()
    custom_data = old_custom_data.copy()
    project = version.requirement.project
    custom_columns = project.get_custom_columns()

    for column in custom_columns:
        value = request.form.get(f'custom_{column}', '').strip()
        old_value = custom_data.get(column, '')
        if old_value != value:
            changes[f'custom_{column}'] = f"{old_value or '–'} → {value or '–'}"
        custom_data[column] = value

    # Handle quantifiable checkbox
    old_quantifiable = old_custom_data.get('is_quantifiable', 'false')
    is_quantifiable = request.form.get('is_quantifiable') == 'on'
    new_quantifiable = 'true' if is_quantifiable else 'false'
    if old_quantifiable != new_quantifiable:
        changes['is_quantifiable'] = (
            f"{'Ja' if old_quantifiable == 'true' else 'Nein'} → {'Ja' if is_quantifiable else 'Nein'}"
        )
    custom_data['is_quantifiable'] = new_quantifiable

    allowed_status = ['Entwurf', 'In Bearbeitung', 'Freigabe', 'Verworfen']
    status_from_form = request.form.get('status')
    status_to_use = status_from_form if status_from_form in allowed_status else version.status
    status_change = None
    if status_to_use != version.status:
        status_change = f"{version.status} → {status_to_use}"
        changes['status'] = status_change

    if not changes:
        flash("Keine Änderungen erkannt. Nimm Änderungen vor, bevor du speicherst.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )

    if status_change and len(changes) == 1:
        version.status = status_to_use
        version.last_modified_by_id = current_user.id
        history_entry = RequirementVersionHistory(
            version_id=version.id,
            changed_by_id=current_user.id,
            change_type='modified',
            changes=json.dumps({'status': status_change})
        )
        db.session.add(history_entry)
        db.session.commit()

        try:
            notify_requirement_updated(version, current_user)
        except Exception:
            pass

        flash(f"Status aktualisiert: {version.status}", "success")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )

    # Determine next version index/label
    latest_version = version.requirement.get_latest_version()
    next_index = (latest_version.version_index + 1) if latest_version else 1
    next_label = chr(ord('A') + (next_index - 1))

    # Create a new version instead of modifying the existing one
    new_version = RequirementVersion(
        requirement_id=version.requirement_id,
        version_index=next_index,
        version_label=next_label,
        title=title,
        description=description,
        category=category,
        status=status_to_use,
        revision=None,
        created_by_id=current_user.id,
        last_modified_by_id=current_user.id
    )
    new_version.set_custom_data(custom_data)
    db.session.add(new_version)
    db.session.flush()

    history_entry = RequirementVersionHistory(
        version_id=new_version.id,
        changed_by_id=current_user.id,
        change_type='created',
        changes=json.dumps({
            'action': 'Neue Version erstellt',
            'from_version': version.version_label,
            'to_version': next_label,
            'changes': changes
        })
    )
    db.session.add(history_entry)

    db.session.commit()

    # Create notifications for requirement update
    try:
        notify_requirement_updated(new_version, current_user)
    except Exception:
        # Don't fail the update if notification fails
        pass

    flash(f"Neue Version {next_label} wurde erstellt. Status: {new_version.status}", "success")
    return redirect(
        url_for(
            'main.manage_project',
            project_id=version.requirement.project_id,
            selected_version_id=new_version.id,
        )
    )


@bp.route("/requirement_version/<int:version_id>/revise", methods=['POST'])
@login_required
def revise_requirement_version(version_id):
    """Handle revision updates with revision numbering logic."""
    from .models import RequirementVersionHistory
    import json

    version = RequirementVersion.query.get_or_404(version_id)
    check_version_access(version)

    if version.status == 'Verworfen':
        flash("Diese Anforderung ist verworfen und kann nur noch gelöscht werden.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )

    title = request.form.get('title', '').strip()
    description = request.form.get('description', '').strip()
    category = request.form.get('category', '').strip()
    revision_key = request.form.get('revision_key', '').strip()
    target_round, revision_number = _parse_revision_key(revision_key)
    update_existing = request.form.get('revision_update') == '1'
    if revision_key and not update_existing:
        revision_number = 0

    if not title or not description:
        flash("Title and description are required for a Revision.", "danger")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )

    allowed_status = ['Entwurf', 'In Bearbeitung', 'Freigabe', 'Verworfen']
    status_from_form = request.form.get('status')

    # Revisions are only allowed after a release
    if not version.requirement.has_release_event():
        flash("Diese Anforderung ist nicht freigegeben und kann nicht revidiert werden.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )

    def status_color_for(value):
        mapping = {
            'Entwurf': 'danger',
            'In Bearbeitung': 'warning',
            'Freigabe': 'success',
            'Verworfen': 'dark',
        }
        return mapping.get(value, 'secondary')

    # Collect custom data from the form
    project = version.requirement.project
    custom_columns = project.get_custom_columns()
    old_custom_data = version.get_custom_data()
    custom_data = old_custom_data.copy()
    for column in custom_columns:
        value = request.form.get(f'custom_{column}', '').strip()
        custom_data[column] = value

    is_quantifiable = request.form.get('is_quantifiable') == 'on'
    custom_data['is_quantifiable'] = 'true' if is_quantifiable else 'false'

    # Load all revision history entries once (ordered)
    history_entries = RequirementVersionHistory.query.filter_by(
        version_id=version.id, change_type='revised'
    ).order_by(RequirementVersionHistory.created_at.asc()).all()

    version_labels = [ver.version_label for ver in (version.requirement.versions or [])]
    if not version_labels:
        version_labels = [version.version_label]

    def normalize_round_label(label: str) -> str:
        candidate = (label or '').strip().upper()[:1]
        if not candidate:
            return version_labels[0] if version_labels else 'A'
        if candidate in version_labels:
            return candidate
        return version_labels[0] if version_labels else candidate

    def next_round_label(label: str) -> str:
        if not version_labels:
            return label or 'A'
        if label not in version_labels:
            return version_labels[0]
        idx = version_labels.index(label)
        if idx + 1 < len(version_labels):
            return version_labels[idx + 1]
        return label

    def snapshot_round_and_number(snapshot: dict) -> tuple[str, int]:
        round_label = normalize_round_label(
            snapshot.get('version_label') or _parse_revision_round(snapshot.get('revision_label'))
        )
        number = snapshot.get('revision_number') or _parse_revision_number(snapshot.get('revision_label'))
        try:
            number = int(number)
        except Exception:
            number = 0
        return round_label, number

    def find_revision_entry(target_number, target_round_label):
        target_round_label = normalize_round_label(target_round_label)
        for entry in history_entries:
            try:
                snapshot = (entry.get_changes() or {}).get('revision_snapshot') or {}
                snap_round, snap_number = snapshot_round_and_number(snapshot)
                if snap_number and snap_round == target_round_label and int(snap_number) == int(target_number):
                    return entry, snapshot
            except Exception:
                continue
        return None, None

    def latest_number_for_round(round_label: str) -> int:
        numbers = []
        for entry in history_entries:
            try:
                snapshot = (entry.get_changes() or {}).get('revision_snapshot') or {}
                snap_round, snap_number = snapshot_round_and_number(snapshot)
                if snap_number and snap_round == round_label:
                    numbers.append(int(snap_number))
            except Exception:
                continue
        return max(numbers) if numbers else 0

    def latest_revision_state() -> tuple[str, int]:
        for entry in reversed(history_entries):
            try:
                snapshot = (entry.get_changes() or {}).get('revision_snapshot') or {}
                if snapshot:
                    snap_round, snap_number = snapshot_round_and_number(snapshot)
                    if snap_number:
                        return snap_round, snap_number
            except Exception:
                continue
        return normalize_round_label(version_labels[0] if version_labels else 'A'), 0

    max_per_round = 3
    last_round, last_number = latest_revision_state()
    if last_number >= max_per_round:
        round_for_new = next_round_label(last_round)
        next_revision_number = 1
    else:
        round_for_new = last_round
        next_revision_number = last_number + 1 if last_number else 1

    def build_snapshot(target_status, target_revision_number, round_label):
        return {
            "title": title,
            "description": description,
            "category": category,
            "status": target_status,
            "status_color": status_color_for(target_status),
            "custom_data": custom_data,
            "is_quantifiable": is_quantifiable,
            "version_label": round_label,
            "revision_label": _format_revision_label(target_revision_number) if target_revision_number else _normalize_revision_label(version.revision),
            "revision_number": target_revision_number if target_revision_number else _parse_revision_number(version.revision),
        }

    # CASE 1: Update an existing revision (revision_number provided)
    if revision_number:
        target_round = normalize_round_label(target_round or round_for_new)
        apply_to_version = target_round == version.version_label
        target_entry, existing_snapshot = find_revision_entry(revision_number, target_round)
        if not target_entry:
            # If the requested revision doesn't exist yet, fall back to creating a new one
            revision_number = None
            round_for_new = target_round
        else:
            prev_custom = existing_snapshot.get('custom_data') or {}
            prev_status = existing_snapshot.get('status') or version.status

            target_status = status_from_form if status_from_form in allowed_status else prev_status

            # Detect changes relative to the selected revision snapshot
            changes_detected = False
            def changed(old_val, new_val):
                return (old_val or "") != (new_val or "")

            if changed(existing_snapshot.get('title'), title):
                changes_detected = True
            if changed(existing_snapshot.get('description'), description):
                changes_detected = True
            if changed(existing_snapshot.get('category'), category):
                changes_detected = True
            for column in custom_columns:
                if changed(prev_custom.get(column, ''), custom_data.get(column, '')):
                    changes_detected = True
            if changed(prev_status, target_status):
                changes_detected = True
            if changed(
                str(existing_snapshot.get('is_quantifiable', False)),
                str(is_quantifiable),
            ):
                changes_detected = True

            if not changes_detected:
                flash("Keine Änderungen erkannt. Nimm Änderungen vor, bevor du revidierst.", "warning")
                return redirect(
                    url_for(
                        'main.manage_project',
                        project_id=version.requirement.project_id,
                        selected_version_id=version.id,
                    )
                )

            new_snapshot = build_snapshot(target_status, revision_number, target_round)
            changes = target_entry.get_changes() or {}
            changes['revision_snapshot'] = new_snapshot
            changes['revision'] = f"Revision {_format_revision_label(revision_number)} aktualisiert"
            target_entry.changes = json.dumps(changes)

            # Only push the edited values onto the live version when editing the latest revision
            latest_number_for_target = latest_number_for_round(target_round)
            if apply_to_version and revision_number >= latest_number_for_target:
                version.title = title
                version.description = description
                version.category = category
                version.last_modified_by_id = current_user.id
                version.set_custom_data(custom_data)
                if target_status in allowed_status:
                    version.status = target_status
                version.revision = _format_revision_label(revision_number)

            db.session.commit()

            flash(
                f"Revision {_format_revision_label(revision_number)} wurde aktualisiert.",
                "success",
            )
            return redirect(
                url_for(
                    'main.manage_project',
                    project_id=version.requirement.project_id,
                    selected_version_id=version.id,
                    selected_req_id=version.requirement_id,
                    selected_revision_key=f"{target_round}-{revision_number}",
                )
            )

    # CASE 2: Create a new revision (existing behavior)
    # Preserve old state for baseline snapshot (before first revision of a round)
    baseline_source = None
    for ver in version.requirement.versions or []:
        if ver.version_label == round_for_new:
            baseline_source = ver
            break
    if not baseline_source:
        baseline_source = version

    old_state_snapshot = {
        "title": baseline_source.title,
        "description": baseline_source.description,
        "category": baseline_source.category,
        "status": baseline_source.status,
        "status_color": baseline_source.get_status_color(),
        "custom_data": baseline_source.get_custom_data(),
        "is_quantifiable": baseline_source.get_custom_data().get('is_quantifiable') in ['true', True],
        "version_label": round_for_new,
        "revision_label": _normalize_revision_label(baseline_source.revision),
        "revision_number": _parse_revision_number(baseline_source.revision),
    }

    changes = {}
    if version.title != title:
        changes['title'] = f"{version.title} → {title}"
    if version.description != description:
        changes['description'] = "Beschreibung geändert"
    if (version.category or '') != category:
        changes['category'] = f"{version.category or '–'} → {category or '–'}"

    for column in custom_columns:
        old_value = old_custom_data.get(column, '')
        new_value = custom_data.get(column, '')
        if (old_value or '') != (new_value or ''):
            changes[f'custom_{column}'] = f"{old_value or '–'} → {new_value or '–'}"

    old_quantifiable = old_custom_data.get('is_quantifiable', 'false')
    new_quantifiable = custom_data.get('is_quantifiable', 'false')
    if old_quantifiable != new_quantifiable:
        changes['is_quantifiable'] = f"{'Ja' if old_quantifiable == 'true' else 'Nein'} → {'Ja' if new_quantifiable == 'true' else 'Nein'}"

    old_status = version.status
    pending_status = None
    if status_from_form and status_from_form in allowed_status:
        if old_status != status_from_form:
            changes['status'] = f"{old_status} → {status_from_form}"
        snapshot_status = status_from_form
        pending_status = status_from_form
    elif changes:
        if old_status != 'In Bearbeitung':
            changes['status'] = f"{old_status} → In Bearbeitung"
        snapshot_status = 'In Bearbeitung'
        pending_status = 'In Bearbeitung'
    else:
        snapshot_status = old_status

    if not changes:
        flash("Keine Änderungen erkannt. Nimm Änderungen vor, bevor du revidierst.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )

    # Recalculate the next revision slot for the chosen round
    latest_for_round = latest_number_for_round(round_for_new)
    if latest_for_round >= max_per_round:
        round_for_new = next_round_label(round_for_new)
        latest_for_round = latest_number_for_round(round_for_new)
    next_revision_number = latest_for_round + 1

    apply_to_version = round_for_new == version.version_label

    # Persist field updates for the latest revision path
    if apply_to_version:
        version.title = title
        version.description = description
        version.category = category
        version.last_modified_by_id = current_user.id
        version.set_custom_data(custom_data)
        if pending_status in allowed_status:
            version.status = pending_status
        version.revision = _format_revision_label(next_revision_number)

    snapshot = build_snapshot(snapshot_status, next_revision_number, round_for_new)

    if next_revision_number == 1:
        changes['revision_baseline'] = old_state_snapshot

    changes['revision'] = f"Revision {version.revision} auf Version {round_for_new} gesetzt"
    changes['revision_snapshot'] = snapshot
    history_entry = RequirementVersionHistory(
        version_id=version.id,
        changed_by_id=current_user.id,
        change_type='revised',
        changes=json.dumps(changes)
    )
    db.session.add(history_entry)

    db.session.commit()

    flash(
        f"Revision {version.revision} wurde auf Version {round_for_new} gesetzt.",
        "success",
    )
    return redirect(
        url_for(
            'main.manage_project',
            project_id=version.requirement.project_id,
            selected_version_id=version.id,
            selected_req_id=version.requirement_id,
            selected_revision_key=f"{round_for_new}-{next_revision_number}",
        )
    )

# Route to toggle quantifiable status
@bp.route("/requirement_version/<int:version_id>/toggle_quantifiable", methods=['POST'])
@login_required
def toggle_quantifiable(version_id):
    from .models import RequirementVersionHistory
    import json
    
    version = RequirementVersion.query.get_or_404(version_id)
    check_version_access(version)

    if version.status == 'Verworfen':
        flash("Diese Anforderung ist verworfen und kann nur noch gelöscht werden.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=version.requirement.project_id,
                selected_version_id=version.id,
            )
        )
    
    custom_data = version.get_custom_data()
    current_value = custom_data.get('is_quantifiable', 'false')
    old_value = 'Ja' if (current_value == 'true' or current_value is True) else 'Nein'
    
    # Toggle value
    if current_value == 'true' or current_value is True:
        custom_data['is_quantifiable'] = 'false'
        new_value = 'Nein'
    else:
        custom_data['is_quantifiable'] = 'true'
        new_value = 'Ja'
    
    version.set_custom_data(custom_data)
    version.last_modified_by_id = current_user.id
    # Status auf 'In Bearbeitung' setzen
    version.status = 'In Bearbeitung'
    
    # Create history entry
    history_entry = RequirementVersionHistory(
        version_id=version.id,
        changed_by_id=current_user.id,
        change_type='modified',
        changes=json.dumps({'is_quantifiable': f"{old_value} → {new_value}"})
    )
    db.session.add(history_entry)
    db.session.commit()
    
    return redirect(
        url_for(
            'main.manage_project',
            project_id=version.requirement.project_id,
            selected_version_id=version.id,
        )
    )

# Route to delete a specific version of a requirement
@bp.route("/requirement_version/<int:version_id>/delete", methods=['POST'])
@login_required
def delete_requirement_version(version_id):
    version = RequirementVersion.query.get_or_404(version_id)
    req = version.requirement

    # Authorization check
    check_requirement_access(req)

    if version.status != 'Verworfen':
        flash("Löschen ist nur möglich, wenn die Anforderung verworfen ist.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=req.project_id,
                selected_version_id=version.id,
            )
        )

    # When a requirement is deleted after being rejected, move the entire requirement to trash
    # (all versions/revisions are treated as deleted together).
    req.is_deleted = True
    flash("Anforderung wurde in den Papierkorb verschoben.", "success")

    db.session.commit()

    return redirect(url_for('main.deleted_requirements_overview'))

# Route to soft delete a requirement (kept for compatibility, but marks all versions as deleted)
@bp.route("/requirement/<int:req_id>/delete", methods=['POST'])
@login_required
def delete_requirement(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    check_requirement_access(req)

    latest_version = req.get_latest_version()
    if latest_version and latest_version.status != 'Verworfen':
        flash("Löschen ist nur möglich, wenn die Anforderung verworfen ist.", "warning")
        return redirect(
            url_for(
                'main.manage_project',
                project_id=req.project_id,
                selected_version_id=latest_version.id,
            )
        )

    # Soft delete
    req.is_deleted = True
    db.session.commit()

    flash("Requirement moved to trash.", "success")
    return redirect(url_for('main.deleted_requirements_overview'))

# Route to restore a deleted requirement
@bp.route("/requirement/<int:req_id>/restore", methods=['POST'])
@login_required
def restore_requirement(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    check_requirement_access(req)

    flash("Wiederherstellen ist deaktiviert. Geloschte Anforderungen konnen nur endgultig entfernt werden.", "warning")
    return redirect(url_for('main.deleted_requirements_overview'))

# Route to permanently delete a requirement
@bp.route("/requirement/<int:req_id>/delete_permanently", methods=['POST'])
@login_required
def delete_requirement_permanently(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    check_requirement_access(req)
    
    project_id = req.project_id
    
    # Permanently delete (cascade will delete all versions)
    db.session.delete(req)
    db.session.commit()
    
    flash("Requirement permanently deleted.", "success")
    return redirect(url_for('main.deleted_requirements_overview'))

# Route to regenerate a single requirement with AI
@bp.route("/requirement/<int:req_id>/regenerate", methods=['POST'])
@login_required
def regenerate_requirement(req_id):
    req = Requirement.query.get_or_404(req_id)
    # Authorization check
    check_requirement_access(req)
    
    # Get the latest version to use as context
    latest_version = req.get_latest_version()
    if not latest_version:
        flash("No existing version found to regenerate.", "danger")
        return redirect(url_for('main.manage_project', project_id=req.project_id))

    if latest_version.status == 'Verworfen':
        flash("Diese Anforderung ist verworfen und kann nur noch gelöscht werden.", "warning")
        return redirect(url_for('main.manage_project', project_id=req.project_id))
    
    try:
        # Get project's custom columns
        custom_columns = req.project.get_custom_columns()
        
        # Prepare context for AI
        context = {
            "project_name": req.project.name,
            "requirement_title": latest_version.title,
            "requirement_description": latest_version.description,
            "requirement_category": latest_version.category or "",
            "custom_data": latest_version.get_custom_data()
        }
        
        # Build complete columns list: title, description, custom columns, category
        columns = ["title", "description"] + custom_columns + ["category"]
        
        # Generate a new version with AI
        result = generate_single_requirement_alternative(context, columns)
        
        if not result:
            flash("Failed to generate alternative. AI returned empty result.", "danger")
            return redirect(url_for('main.manage_project', project_id=req.project_id))
        
        # Calculate next version
        next_index = latest_version.version_index + 1
        next_label = chr(ord('A') + (next_index - 1))
        
        # Create new version
        new_version = RequirementVersion(
            requirement_id=req.id,
            version_index=next_index,
            version_label=next_label,
            title=result.get("title", latest_version.title),
            description=result.get("description", latest_version.description),
            category=result.get("category", latest_version.category),
            status="Entwurf",  # New version starts as "Entwurf"
            revision=None,
            created_by_id=current_user.id  # Track who created this version
        )
        
        # Get custom data from AI result or copy from previous version
        custom_data = {}
        for col in custom_columns:
            # Try to get value from AI result first, fallback to previous version
            value = result.get(col, latest_version.get_custom_data().get(col, ""))
            if value:
                custom_data[col] = value
        
        if custom_data:
            new_version.set_custom_data(custom_data)
        
        db.session.add(new_version)
        db.session.flush()  # Get the ID for history entry
        
        # Create history entry for regeneration
        from .models import RequirementVersionHistory
        import json
        history_entry = RequirementVersionHistory(
            version_id=new_version.id,
            changed_by_id=current_user.id,
            change_type='created',
            changes=json.dumps({'action': 'Version regeneriert (KI)', 'version': next_label})
        )
        db.session.add(history_entry)
        db.session.commit()
        
        flash(f"New version {next_label} generated successfully!", "success")
        
    except Exception as e:
        flash(f"Error generating alternative: {str(e)}", "danger")
    
    return redirect(url_for('main.manage_project', project_id=req.project_id))

def generate_single_requirement_alternative(context, columns):
    """Generate an alternative version of a requirement using AI."""
    try:
        # Prepare prompt for AI
        prompt = f"""
        Generate an alternative version of the following requirement:
        
        Project: {context['project_name']}
        
        Original Requirement:
        Title: {context['requirement_title']}
        Description: {context['requirement_description']}
        Category: {context['requirement_category']}
        
        Additional Context:
        {context['custom_data']}
        
        Please provide an improved version with:
        1. A clearer title
        2. A more detailed description
        3. The same or improved category
        
        Keep the core meaning but enhance clarity, completeness, and precision.
        """
        
        # Call the AI service
        ai_result = generate_requirements(prompt, {}, columns)
        
        # We expect a list of requirements, but we only need the first one
        if ai_result and len(ai_result) > 0:
            return ai_result[0]
        
        return None
        
    except Exception as e:
        print(f"Error in generate_single_requirement_alternative: {str(e)}")
        raise

# Route to export project requirements to Excel
@bp.route("/project/<int:project_id>/export_excel", methods=["GET", "POST"])
@login_required
def export_excel(project_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    from flask import send_file
    from .models import RequirementVersionHistory
    
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    
    # Get all non-deleted requirements with their latest versions
    requirements = Requirement.query.filter_by(
        project_id=project_id,
        is_deleted=False
    ).all()
    
    # Get custom columns
    custom_columns = project.get_custom_columns()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    
    # Define headers in requested order
    headers = [
        "Verantwortlicher",
        "Revision",
        "Version",
        "ID",
        "Anforderung",
        "Beschreibung",
    ] + custom_columns + ["Kategorie", "Status"]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="top")
    
    selection_payload = request.form.get("selection_payload", "").strip()
    selected_items = []
    if selection_payload:
        try:
            selected_items = json.loads(selection_payload)
        except Exception:
            selected_items = []

    def _status_is_release(value: str) -> bool:
        return str(value or "").lower() == "freigabe"

    def _find_revision_snapshot(revision_version_id, revision_key_value):
        if not revision_version_id or not revision_key_value:
            return None
        target_round, target_number = _parse_revision_key(revision_key_value)
        if not target_round or not target_number:
            return None
        entries = (
            RequirementVersionHistory.query
            .filter_by(version_id=revision_version_id, change_type='revised')
            .order_by(RequirementVersionHistory.created_at.asc())
            .all()
        )
        for entry in entries:
            try:
                changes = entry.get_changes() or {}
            except Exception:
                changes = {}
            snapshot = changes.get('revision_snapshot') or {}
            if not snapshot:
                continue
            snap_round = snapshot.get('version_label') or _parse_revision_round(snapshot.get('revision_label'))
            snap_number = snapshot.get('revision_number') or _parse_revision_number(snapshot.get('revision_label'))
            try:
                snap_number = int(snap_number)
            except Exception:
                snap_number = 0
            if snap_round and str(snap_round).upper()[:1] == target_round and snap_number == int(target_number):
                return snapshot
        return None

    # Write data rows
    row_num = 2
    display_id = 1

    if selected_items:
        for item in selected_items:
            try:
                req_id = int(item.get("req_id"))
            except Exception:
                continue
            req = Requirement.query.filter_by(
                id=req_id,
                project_id=project_id,
                is_deleted=False,
            ).first()
            if not req:
                continue

            version_id = item.get("version_id")
            revision_key_value = item.get("revision_key") or ""
            revision_version_id = item.get("revision_version_id") or version_id

            version_obj = None
            if version_id:
                version_obj = RequirementVersion.query.get(version_id)
                if not version_obj or version_obj.requirement_id != req.id:
                    version_obj = None

            snapshot = _find_revision_snapshot(revision_version_id, revision_key_value)
            if snapshot and _status_is_release(snapshot.get("status")):
                custom_data = snapshot.get("custom_data") or {}
                creator = (
                    version_obj.created_by.email
                    if version_obj and version_obj.created_by
                    else "–"
                )
                revision = _normalize_revision_label(snapshot.get("revision_label"))
                row_data = [
                    creator,
                    revision,
                    snapshot.get("version_label") or (version_obj.version_label if version_obj else ""),
                    display_id,
                    snapshot.get("title") or "–",
                    snapshot.get("description") or "–",
                ]
                for col in custom_columns:
                    row_data.append(custom_data.get(col, "–"))
                row_data.append(snapshot.get("category") or "–")
                row_data.append(snapshot.get("status") or "–")
            elif version_obj and _status_is_release(version_obj.status):
                custom_data = version_obj.get_custom_data()
                creator = version_obj.created_by.email if version_obj.created_by else "–"
                revision = version_obj.get_revision_display() or ""
                row_data = [
                    creator,
                    revision,
                    version_obj.version_label,
                    display_id,
                    version_obj.title,
                    version_obj.description,
                ]
                for col in custom_columns:
                    row_data.append(custom_data.get(col, "–"))
                row_data.append(version_obj.category or "–")
                row_data.append(version_obj.status)
            else:
                continue

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            row_num += 1
            display_id += 1
    else:
        for req in requirements:
            latest_version = req.get_latest_version()
            if not latest_version:
                continue
            latest_released = req.get_latest_released_version()
            if not latest_released:
                continue
            if latest_released.status != 'Freigabe':
                continue

            custom_data = latest_released.get_custom_data()

            # Prepare row data in requested order
            creator = latest_released.created_by.email if latest_released.created_by else "–"
            revision = latest_released.get_revision_display() or ""
            row_data = [
                creator,
                revision,
                latest_released.version_label,
                display_id,
                latest_released.title,
                latest_released.description,
            ]

            # Add custom column values
            for col in custom_columns:
                row_data.append(custom_data.get(col, "–"))

            # Add category and status
            row_data.append(latest_released.category or "–")
            row_data.append(latest_released.status)

            # Write row
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            row_num += 1
            display_id += 1
    
    # Set column widths (readable layout)
    ws.column_dimensions['A'].width = 24  # Verantwortlicher
    ws.column_dimensions['B'].width = 12  # Revision
    ws.column_dimensions['C'].width = 10  # Version
    ws.column_dimensions['D'].width = 8   # ID
    ws.column_dimensions['E'].width = 35  # Anforderung
    ws.column_dimensions['F'].width = 60  # Beschreibung

    # Set widths for custom columns starting at column G
    col_letter_start = ord('G')
    for i, col in enumerate(custom_columns):
        col_letter = chr(col_letter_start + i)
        ws.column_dimensions[col_letter].width = 22

    # Set widths for category and status after custom columns
    col_letter = chr(col_letter_start + len(custom_columns))
    ws.column_dimensions[col_letter].width = 20  # Kategorie
    col_letter = chr(col_letter_start + len(custom_columns) + 1)
    ws.column_dimensions[col_letter].width = 15  # Status
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create filename
    filename = f"requirements_{project.name.replace(' ', '_')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route("/project/<int:project_id>/export_sysml", methods=["GET", "POST"])
@login_required
def export_sysml(project_id):
    from io import BytesIO
    import tempfile
    import pandas as pd
    from app.utils.sysml_v2_generator import generate_sysmlv2_code

    project = Project.query.get_or_404(project_id)
    check_project_access(project)

    requirements = Requirement.query.filter_by(
        project_id=project_id,
        is_deleted=False
    ).all()

    requirements_data = []
    display_id = 1
    for req in requirements:
        latest_version = req.get_latest_version()
        if not latest_version:
            continue

        responsible = latest_version.created_by.email if latest_version.created_by else "-"
        revision = latest_version.revision or "Entwurf"
        version = latest_version.version_label or ""
        category = latest_version.category or ""
        status = latest_version.status or "Entwurf"

        requirements_data.append({
            "Verantwortlicher": responsible,
            "Revision": revision,
            "Version": version,
            "ID": display_id,
            "Anforderung": latest_version.title,
            "Beschreibung": latest_version.description,
            "Kategorie": category,
            "Status": status,
        })
        display_id += 1

    df = pd.DataFrame(requirements_data)

    with tempfile.NamedTemporaryFile(mode="w+", suffix=".txt", delete=False, encoding="utf-8") as tmp_file:
        tmp_path = tmp_file.name

    generate_sysmlv2_code(df, tmp_path)

    with open(tmp_path, "rb") as tmp_file:
        output = BytesIO(tmp_file.read())
    output.seek(0)

    filename = f"requirements_{project.name.replace(' ', '_')}.sysml.txt"

    return send_file(
        output,
        mimetype="text/plain",
        as_attachment=True,
        download_name=filename
    )

# Route to import requirements from Excel
@bp.route("/project/<int:project_id>/import_excel", methods=['POST'])
@login_required
def import_excel(project_id):
    from openpyxl import load_workbook
    from werkzeug.utils import secure_filename
    import os
    
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    
    # Check if file was uploaded
    if 'excel_file' not in request.files:
        flash("Keine Datei ausgewählt.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    file = request.files['excel_file']
    
    if file.filename == '':
        flash("Keine Datei ausgewählt.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash("Bitte laden Sie eine Excel-Datei (.xlsx oder .xls) hoch.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    try:
        # Load workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        # Get custom columns for this project
        custom_columns = project.get_custom_columns()
        
        # Read header row to map columns
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Find column indices
        title_idx = None
        description_idx = None
        category_idx = None
        status_idx = None
        custom_col_indices = {}
        
        for idx, header in enumerate(headers):
            header_lower = header.lower()
            if header_lower in ['title', 'titel']:
                title_idx = idx
            elif header_lower in ['description', 'beschreibung']:
                description_idx = idx
            elif header_lower in ['category', 'kategorie']:
                category_idx = idx
            elif header_lower in ['status']:
                status_idx = idx
            elif header in custom_columns:
                custom_col_indices[header] = idx
        
        if title_idx is None or description_idx is None:
            flash("Excel-Datei muss mindestens 'Title' und 'Beschreibung' Spalten enthalten.", "danger")
            return redirect(url_for('main.manage_project', project_id=project_id))
        
        # Import rows (skip header)
        imported_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= title_idx:
                continue
            
            title = row[title_idx]
            if not title or str(title).strip() == '':
                continue
            
            title = str(title).strip()
            description = str(row[description_idx]).strip() if description_idx < len(row) and row[description_idx] else ""
            
            if not description:
                continue
            
            category = str(row[category_idx]).strip() if category_idx is not None and category_idx < len(row) and row[category_idx] else ""
            status = str(row[status_idx]).strip() if status_idx is not None and status_idx < len(row) and row[status_idx] else "Entwurf"
            
            # Validate status
            if status not in ['Entwurf', 'In Bearbeitung', 'Freigabe']:
                status = 'Entwurf'
            
            # Create requirement
            from .agent import normalize_key
            key = normalize_key(title)
            
            req = Requirement.query.filter_by(project_id=project_id, key=key).first()
            
            if not req:
                req = Requirement(project_id=project_id, key=key)
                db.session.add(req)
                db.session.flush()
                version_index = 1
                version_label = 'A'
            else:
                # Create new version
                last_version = req.versions[-1] if req.versions else None
                version_index = last_version.version_index + 1 if last_version else 1
                version_label = chr(ord('A') + (version_index - 1))
            
            # Create version
            new_version = RequirementVersion(
                requirement_id=req.id,
                version_index=version_index,
                version_label=version_label,
                title=title,
                description=description,
                category=category,
                status=status,
                created_by_id=current_user.id
            )
            
            # Add custom column data
            custom_data = {}
            for col_name, col_idx in custom_col_indices.items():
                if col_idx < len(row) and row[col_idx]:
                    custom_data[col_name] = str(row[col_idx]).strip()
            
            if custom_data:
                new_version.set_custom_data(custom_data)
            
            db.session.add(new_version)
            db.session.flush()  # Get the ID for history entry
            
            # Create history entry for import
            from .models import RequirementVersionHistory
            import json
            history_entry = RequirementVersionHistory(
                version_id=new_version.id,
                changed_by_id=current_user.id,
                change_type='created',
                changes=json.dumps({'action': 'Version importiert (Excel)', 'version': version_label})
            )
            db.session.add(history_entry)
            
            imported_count += 1
        
        db.session.commit()
        flash(f"{imported_count} Anforderungen erfolgreich importiert!", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Fehler beim Importieren: {str(e)}", "danger")
    
    return redirect(url_for('main.manage_project', project_id=project_id))

# Route to share project with another user
@bp.route("/project/<int:project_id>/share", methods=['POST'])
@login_required
def share_project(project_id):
    from .models import User
    
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    
    email = request.form.get('email', '').strip()
    if not email:
        flash("Bitte geben Sie eine E-Mail-Adresse ein.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    # Find user by email
    user = User.query.filter_by(email=email).first()
    if not user:
        flash(f"Benutzer mit E-Mail '{email}' nicht gefunden.", "danger")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    if user.id == current_user.id:
        flash("Sie können das Projekt nicht mit sich selbst teilen.", "warning")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    # Check if already shared
    if user in project.shared_with:
        flash(f"Projekt ist bereits mit {email} geteilt.", "warning")
        return redirect(url_for('main.manage_project', project_id=project_id))
    
    # Share project
    project.shared_with.append(user)
    db.session.commit()
    
    flash(f"Projekt erfolgreich mit {email} geteilt!", "success")
    return redirect(url_for('main.manage_project', project_id=project_id))

# Route to unshare project
@bp.route("/project/<int:project_id>/unshare/<int:user_id>", methods=['POST'])
@login_required
def unshare_project(project_id, user_id):
    from .models import User
    
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
    
    user = User.query.get_or_404(user_id)
    
    if user in project.shared_with:
        project.shared_with.remove(user)
        db.session.commit()
        flash(f"Projekt-Freigabe für {user.email} entfernt.", "success")
    else:
        flash("Benutzer hat keinen Zugriff auf dieses Projekt.", "warning")
    
    return redirect(url_for('main.manage_project', project_id=project_id))

# Route to toggle requirement version block status
@bp.route("/requirement_version/<int:version_id>/toggle_block", methods=['POST'])
@login_required
def toggle_block_requirement(version_id):
    from datetime import datetime
    
    version = RequirementVersion.query.get_or_404(version_id)
    project = version.requirement.project
    
    # Authorization check - only project owner can block
    if project.user_id != current_user.id:
        abort(403)
    
    # Toggle block status
    if version.is_blocked:
        # Unblock
        version.is_blocked = False
        version.blocked_by_id = None
        version.blocked_at = None
        flash(f"Version {version.version_label} wurde freigegeben.", "success")
    else:
        # Block
        version.is_blocked = True
        version.blocked_by_id = current_user.id
        version.blocked_at = datetime.utcnow()
        flash(f"Version {version.version_label} wurde blockiert.", "warning")
    
    db.session.commit()
    return redirect(
        url_for(
            'main.manage_project',
            project_id=project.id,
            selected_version_id=version.id,
        )
    )


@bp.route("/project/<int:project_id>/detect_conflicts")
@login_required
def detect_conflicts_route(project_id):
    from .services.ai_client import detect_conflicts
    
    project = Project.query.get_or_404(project_id)
    check_project_access(project)
        
    # Gather all latest versions of not deleted requirements
    # Similar logic to export/view
    requirements = Requirement.query.filter_by(
        project_id=project_id,
        is_deleted=False
    ).all()
    
    req_list = []
    display_id = 1
    
    for req in requirements:
        latest = req.get_latest_version()
        if latest:
            req_data = {
                "id": display_id, # User friendly ID
                "db_id": req.id,
                "title": latest.title,
                "description": latest.description
            }
            req_list.append(req_data)
            display_id += 1
            
    if len(req_list) < 2:
        return jsonify({'conflicts': []}) # Need at least 2 to have a conflict
        
    conflicts = detect_conflicts(req_list)
    
    return jsonify({'conflicts': conflicts})


@bp.route("/requirement_version/<int:version_id>/generate_tests", methods=['POST'])
@login_required
def generate_tests_route(version_id):
    from .services.ai_client import generate_test_cases
    
    version = RequirementVersion.query.get_or_404(version_id)
    try:
        check_version_access(version)
    except:
        return jsonify({'error': 'Zugriff verweigert'}), 403
        
    test_cases = generate_test_cases(version.title, version.description)
    return jsonify({'result': test_cases})


# ============================================================================
# Import Helper Functions for Comments and Notifications
# ============================================================================
from .utils.notifications import (
    notify_requirement_updated,
    notify_requirement_created,
    notify_comment_added,
    parse_mentions,
    find_user_by_mention
)

# ============================================================================
# Comment Routes
# ============================================================================

@bp.route("/requirement_version/<int:version_id>/comments", methods=['GET'])
@login_required
def get_comments(version_id):
    """Get all comments for a requirement version."""
    version = RequirementVersion.query.get_or_404(version_id)
    check_version_access(version)
    
    # Get all non-deleted comments, ordered by creation date
    comments = RequirementComment.query.filter_by(
        version_id=version_id,
        is_deleted=False
    ).order_by(RequirementComment.created_at.asc()).all()
    
    # Build comment tree
    comments_data = []
    for comment in comments:
        comment_data = {
            'id': comment.id,
            'text': comment.text,
            'author': {
                'id': comment.author.id,
                'email': comment.author.email,
                'name': comment.author.email.split('@')[0]
            },
            'created_at': comment.created_at.isoformat(),
            'updated_at': comment.updated_at.isoformat(),
            'parent_comment_id': comment.parent_comment_id,
            'replies': []
        }
        comments_data.append(comment_data)
    
    # Build tree structure
    comment_dict = {c['id']: c for c in comments_data}
    root_comments = []
    for comment in comments_data:
        if comment['parent_comment_id']:
            parent = comment_dict.get(comment['parent_comment_id'])
            if parent:
                parent['replies'].append(comment)
        else:
            root_comments.append(comment)
    
    return jsonify({'comments': root_comments})

@bp.route("/requirement_version/<int:version_id>/comments", methods=['POST'])
@login_required
def create_comment(version_id):
    """Create a new comment on a requirement version."""
    version = RequirementVersion.query.get_or_404(version_id)
    check_version_access(version)
    
    data = request.get_json()
    text = data.get('text', '').strip()
    parent_comment_id = data.get('parent_comment_id')  # Optional, for replies
    
    if not text:
        return jsonify({'error': 'Kommentar-Text ist erforderlich'}), 400
    
    # Validate parent comment if provided
    if parent_comment_id:
        parent = RequirementComment.query.get(parent_comment_id)
        if not parent or parent.version_id != version_id or parent.is_deleted:
            return jsonify({'error': 'Ungültiger Parent-Kommentar'}), 400
    
    # Create comment
    comment = RequirementComment(
        version_id=version_id,
        author_id=current_user.id,
        text=text,
        parent_comment_id=parent_comment_id
    )
    db.session.add(comment)
    db.session.commit()
    
    # Create notifications
    try:
        notify_comment_added(comment, current_user)
    except Exception as e:
        # Don't fail comment creation if notification fails
        pass
    
    # Return created comment
    return jsonify({
        'id': comment.id,
        'text': comment.text,
        'author': {
            'id': comment.author.id,
            'email': comment.author.email,
            'name': comment.author.email.split('@')[0]
        },
        'created_at': comment.created_at.isoformat(),
        'updated_at': comment.updated_at.isoformat(),
        'parent_comment_id': comment.parent_comment_id
    }), 201

@bp.route("/comment/<int:comment_id>", methods=['PUT'])
@login_required
def update_comment(comment_id):
    """Update an existing comment."""
    comment = RequirementComment.query.get_or_404(comment_id)
    check_version_access(comment.version)
    
    # Check if user is the author
    if comment.author_id != current_user.id:
        return jsonify({'error': 'Keine Berechtigung zum Bearbeiten dieses Kommentars'}), 403
    
    if comment.is_deleted:
        return jsonify({'error': 'Kommentar wurde gelöscht'}), 400
    
    data = request.get_json()
    text = data.get('text', '').strip()
    
    if not text:
        return jsonify({'error': 'Kommentar-Text ist erforderlich'}), 400
    
    comment.text = text
    db.session.commit()
    
    return jsonify({
        'id': comment.id,
        'text': comment.text,
        'updated_at': comment.updated_at.isoformat()
    })

@bp.route("/comment/<int:comment_id>", methods=['DELETE'])
@login_required
def delete_comment(comment_id):
    """Delete (soft delete) a comment."""
    comment = RequirementComment.query.get_or_404(comment_id)
    check_version_access(comment.version)
    
    # Check if user is the author or project owner
    project = comment.version.requirement.project
    if comment.author_id != current_user.id and project.user_id != current_user.id:
        return jsonify({'error': 'Keine Berechtigung zum Löschen dieses Kommentars'}), 403
    
    # Soft delete
    comment.is_deleted = True
    db.session.commit()
    
    return jsonify({'success': True})

# ============================================================================
# Notification Routes
# ============================================================================

@bp.route("/notifications", methods=['GET'])
@login_required
def get_notifications():
    """Get all notifications for current user."""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    unread_only = request.args.get('unread_only', 'false').lower() == 'true'
    
    query = Notification.query.filter_by(user_id=current_user.id)
    if unread_only:
        query = query.filter_by(is_read=False)
    
    query = query.order_by(Notification.created_at.desc())
    
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)
    
    notifications = []
    for notif in paginated.items:
        metadata = notif.get_metadata()
        notifications.append({
            'id': notif.id,
            'type': notif.notification_type,
            'title': notif.title,
            'message': notif.message,
            'is_read': notif.is_read,
            'created_at': notif.created_at.isoformat(),
            'related_type': notif.related_type,
            'related_id': notif.related_id,
            'metadata': metadata
        })
    
    return jsonify({
        'notifications': notifications,
        'total': paginated.total,
        'page': page,
        'pages': paginated.pages,
        'unread_count': Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    })

@bp.route("/notifications/unread_count", methods=['GET'])
@login_required
def get_unread_notification_count():
    """Get count of unread notifications."""
    count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    return jsonify({'unread_count': count})

@bp.route("/notification/<int:notification_id>/read", methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    """Mark a notification as read."""
    notification = Notification.query.get_or_404(notification_id)
    
    if notification.user_id != current_user.id:
        return jsonify({'error': 'Keine Berechtigung'}), 403
    
    notification.mark_as_read()
    db.session.commit()
    
    return jsonify({'success': True})

@bp.route("/notifications/mark_all_read", methods=['POST'])
@login_required
def mark_all_notifications_read():
    """Mark all notifications as read for current user."""
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update(
        {'is_read': True, 'read_at': datetime.utcnow()}
    )
    db.session.commit()
    
    return jsonify({'success': True})

@bp.route("/requirement_version/<int:version_id>/info")
@login_required
def get_version_info(version_id):
    """Get basic info about a requirement version (for navigation)."""
    version = RequirementVersion.query.get_or_404(version_id)
    check_version_access(version)
    
    return jsonify({
        'requirement_id': version.requirement_id,
        'version_id': version.id,
        'project_id': version.requirement.project_id
    })

@bp.route("/hello")
def hello():
    return "Hello from Blüprint!"

@bp.route("/requirement/<int:req_id>/toggle_funktional", methods=["POST"])

@bp.route("/requirement/<int:req_id>/toggle_funktional", methods=["POST"])
@login_required
def toggle_funktional(req_id):
    from .models import RequirementVersionHistory

    req = Requirement.query.get_or_404(req_id)
    check_requirement_access(req)
    latest_version = req.get_latest_version()
    if latest_version and latest_version.status == 'Verworfen':
        flash("Diese Anforderung ist verworfen und kann nur noch gelöscht werden.", "warning")
        return redirect(
            request.referrer or url_for('main.manage_project', project_id=req.project_id)
        )
    # Toggle Wert setzen
    new_value = request.form.get("funktional")
    old_value = req.funktional
    req.funktional = bool(int(new_value))
    latest_version = req.get_latest_version()
    if latest_version and old_value != req.funktional:
        latest_version.last_modified_by_id = current_user.id
        # Status auf 'In Bearbeitung' setzen
        latest_version.status = 'In Bearbeitung'
        history_entry = RequirementVersionHistory(
            version_id=latest_version.id,
            changed_by_id=current_user.id,
            change_type='modified',
            changes=json.dumps({
                'funktional': f"{'Ja' if old_value else 'Nein'} → {'Ja' if req.funktional else 'Nein'}"
            })
        )
        db.session.add(history_entry)
    db.session.commit()
    flash(
        f"Funktionalität für Anforderung #{req.id} wurde "
        f"{'aktiviert' if req.funktional else 'deaktiviert' }.",
        "success"
    )
    return redirect(request.referrer or url_for('main.manage_project', project_id=req.project_id))

@bp.route("/requirement_version/<int:version_id>/analyze", methods=['POST'])
@login_required
def analyze_requirement_route(version_id):
    from .services.ai_client import analyze_requirement

    version = RequirementVersion.query.get_or_404(version_id)
    try:
        check_version_access(version)
    except:
        return jsonify({'error': 'Zugriff verweigert'}), 403

    analysis = analyze_requirement(version.title, version.description, version.status)
    quantifiable = (
        analysis.get("quantifiable_assessment", {}).get("is_quantifiable") is True
    )

    status_value = (version.status or "").strip().lower()
    if status_value in {"freigabe"}:
        status_color = "GRÜN"
    elif status_value in {"in bearbeitung"}:
        status_color = "GELB"
    else:
        status_color = "ROT"

    analysis["status_context"] = {
        "status": version.status,
        "ampel": status_color if quantifiable else "NICHT_ANWENDBAR"
    }
    return jsonify({'result': analysis})

