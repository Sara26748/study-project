from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Anforderungen"

# Define headers
headers = ["Titel", "Beschreibung", "Kategorie", "Status", "Priorität"]

# Write headers with bold font
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Add sample data
data = [
    [
        "Benutzeranmeldung", 
        "Der Benutzer muss sich mit E-Mail und Passwort anmelden können.", 
        "Funktional", 
        "Offen", 
        "Hoch"
    ],
    [
        "Temperaturregelung", 
        "Das System muss die Raumtemperatur automatisch basierend auf Zeitplänen regeln.", 
        "Funktional", 
        "In Arbeit", 
        "Mittel"
    ],
    [
        "Datenverschlüsselung", 
        "Alle Benutzerdaten müssen verschlüsselt gespeichert werden (AES-256).", 
        "Sicherheit", 
        "Offen", 
        "Kritisch"
    ],
    [
        "Lichtsteuerung", 
        "Das System muss erlauben, Lichter per App ein- und auszuschalten.", 
        "Funktional", 
        "Offen", 
        "Niedrig"
    ]
]

# Write data rows
for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Adjust column widths for better readability
ws.column_dimensions['A'].width = 25  # Title
ws.column_dimensions['B'].width = 50  # Description
ws.column_dimensions['C'].width = 15  # Category
ws.column_dimensions['D'].width = 10  # Status
ws.column_dimensions['E'].width = 10  # Priority

# Save the file
filename = "beispiel_anforderungen.xlsx"
wb.save(filename)
print(f"Excel-Datei '{filename}' wurde erfolgreich erstellt.")
